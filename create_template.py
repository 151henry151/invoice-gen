from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_invoice_template():
    wb = Workbook()
    ws = wb.active
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Style definitions
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Company Information (B2-G6)
    ws.merge_cells('B2:G2')
    ws['B2'] = 'Company Logo'
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells('B3:G3')
    ws['B3'] = 'Company Name'
    ws['B3'].font = title_font
    
    ws.merge_cells('B4:G4')
    ws['B4'] = 'Company Address'
    
    ws.merge_cells('B5:G5')
    ws['B5'] = 'Company Email'
    
    ws.merge_cells('B6:G6')
    ws['B6'] = 'Company Phone'
    
    # Invoice Information (G8-G10)
    ws['G8'] = 'Invoice #'
    ws['G8'].font = header_font
    ws['G8'].alignment = Alignment(horizontal='right')
    
    ws['G9'] = 'Date:'
    ws['G9'].font = header_font
    ws['G9'].alignment = Alignment(horizontal='right')
    
    # Client Information (B14-B17)
    ws['B14'] = 'Client Name'
    ws['B14'].font = header_font
    
    ws['B15'] = 'Client Address'
    ws['B15'].font = header_font
    
    ws['B16'] = 'Client Email'
    ws['B16'].font = header_font
    
    ws['B17'] = 'Client Phone'
    ws['B17'].font = header_font
    
    # Line Items Header (B22-G22)
    headers = ['Date', 'Description', 'Quantity', 'Unit Price', 'Total']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=22, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Line Items (B23-G32)
    for row in range(23, 33):
        for col in range(2, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col in [5, 6, 7]:  # Unit Price and Total columns
                cell.number_format = '"$"#,##0.00'
    
    # Notes Section (B34)
    ws['B34'] = 'Notes:'
    ws['B34'].font = header_font
    
    # Totals Section (G39)
    ws['G39'] = 'Total:'
    ws['G39'].font = header_font
    ws['G39'].alignment = Alignment(horizontal='right')
    
    # Save the template
    wb.save('templates/Invoice.xlsx')

if __name__ == '__main__':
    create_invoice_template() 